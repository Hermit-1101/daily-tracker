from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
from datetime import datetime
import os

app = Flask(__name__)

EXCEL_FILE = "timetable.xlsx"
TIME_DATA = [
    '5:00 – 5:15 AM', '5:15 – 5:25 AM', '5:30 – 6:00 AM', '6:00 – 6:15 AM',
    '6:15 – 6:30 AM', '6:30 – 6:45 AM', '6:45 – 7:00 AM', '7:00 – 8:00 AM',
    '8:00 – 8:30 AM', '8:30 – 9:00 AM', '9:00 – 11:00 AM', '11:00 – 11:30 AM',
    '11:30 – 1:30 PM', '1:30 – 2:00 PM', '2:00 – 2:30 PM', '2:30 – 4:30 PM',
    '4:30 – 5:15 PM', '5:30 – 7:30 PM', '7:30 – 8:00 PM', '8:00 – 8:30 PM',
    '8:30 – 9:00 PM', '9:00 – 9:45 PM', '10:00 PM'
]

TASK_DATA = [
    'Wake up, drink water, Wim Hof breathing, freshen up',
    'Light pre-workout snack (banana + peanut butter/nuts)',
    'Dumbbell Workout (4 days/week) or Run (on 2 rest days)',
    'Post-workout meal (high protein)',
    'Cold water bath',
    'Read 10–15 pages of a book',
    'Light mobility drills / stretching',
    'Coding Practice or Project Work',
    'Breakfast + Household chores',
    'Quick revision or prep for online classes',
    'Online Class: Frontend Development',
    'Break + Light snack',
    'Online Class: Java',
    'Lunch (balanced meal)',
    'Power Nap (20–30 mins)',
    'Online Class: Manual Testing',
    'Interview Prep / Aptitude Practice',
    'Online Class: Python',
    'Dinner (light but protein-rich)',
    'Light stretching / Walk + Journaling',
    'No screen time: meditation / breathing / book',
    'Wind down + Sleep prep',
    'Sleep (Non-negotiable 😴)'
]

def init_excel():
    today = datetime.today().strftime('%Y-%m-%d')
    if not os.path.exists(EXCEL_FILE):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Time'
        ws['B1'] = 'Task'
        ws['C1'] = today
        for i in range(len(TIME_DATA)):
            ws.cell(row=i + 2, column=1, value=TIME_DATA[i])
            ws.cell(row=i + 2, column=2, value=TASK_DATA[i])
            ws.cell(row=i + 2, column=3, value="Incomplete")
        wb.save(EXCEL_FILE)
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        last_col = ws.max_column
        last_date = ws.cell(row=1, column=last_col).value
        if last_date != today:
            ws.cell(row=1, column=last_col + 1, value=today)
            for i in range(len(TIME_DATA)):
                ws.cell(row=i + 2, column=last_col + 1, value="Incomplete")
            wb.save(EXCEL_FILE)

@app.route('/')
def index():
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    today = datetime.today().strftime('%Y-%m-%d')
    date_col = ws.max_column

    task_statuses = []
    for i in range(2, len(TIME_DATA) + 2):
        status = ws.cell(row=i, column=date_col).value
        task_statuses.append({
            "time": ws.cell(row=i, column=1).value,
            "task": ws.cell(row=i, column=2).value,
            "status": status == "Complete"
        })

    return render_template('index.html', date=today, tasks=task_statuses)

@app.route('/update', methods=['POST'])
def update():
    data = request.get_json()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    date_col = ws.max_column
    for i, task in enumerate(data['tasks']):
        ws.cell(row=i + 2, column=date_col, value="Complete" if task else "Incomplete")
    wb.save(EXCEL_FILE)
    return jsonify({"success": True})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
